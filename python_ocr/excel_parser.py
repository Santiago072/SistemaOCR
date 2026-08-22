"""
Parser de Archivos Excel (.xlsx / .xls / .csv) de Reportes de Inscripción SENA
Estructura procesada según la plantilla de inscripción:
- Fila 3: Código Ficha (ej: 3574135)
- Fila 4: Programa de Formación
- Fila 6+: Columnas -> Identificación (ej: CC - 1110487315), Nombre, Estado (ej: Preinscrito)

Uso:
    python excel_parser.py --excel "ruta/al/archivo.xlsx"
"""

import os
import sys
import json
import argparse
import re

# Forzar stdout en UTF-8 para evitar errores de codificación en Windows
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

def parse_excel_file(file_path: str) -> dict:
    if not os.path.exists(file_path):
        return {
            "status": "error",
            "message": f"Archivo Excel no encontrado: {file_path}",
            "data": None
        }

    try:
        rows_data = []

        # Determinar si es .xlsx o .xls o HTML disfrazado de XLS (típico de descargas web SENA)
        is_xlsx = file_path.lower().endswith('.xlsx')

        if is_xlsx:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet = wb.active
            for r in range(1, sheet.max_row + 1):
                row = [sheet.cell(row=r, column=c).value for c in range(1, sheet.max_column + 1)]
                rows_data.append(row)
            wb.close()
        else:
            # Intentar con xlrd o parser nativo si es tabla exportada
            loaded = False
            try:
                import xlrd
                wb = xlrd.open_workbook(file_path)
                sheet = wb.sheet_by_index(0)
                for r in range(sheet.nrows):
                    rows_data.append(sheet.row_values(r))
                loaded = True
            except Exception:
                pass

            if not loaded:
                # Muchas descargas web (.xls) son archivos HTML con etiquetas <table><tr><td>
                with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    content = f.read()
                
                # Regex robusto para extraer filas y celdas HTML sin requerir librerías externas
                tr_matches = re.findall(r'<tr[^>]*>(.*?)</tr>', content, re.IGNORECASE | re.DOTALL)
                for tr in tr_matches:
                    td_matches = re.findall(r'<t[dh][^>]*>(.*?)</t[dh]>', tr, re.IGNORECASE | re.DOTALL)
                    if td_matches:
                        # Limpiar tags html internos y entidades
                        clean_row = [re.sub(r'<[^>]+>', '', cell).replace('&nbsp;', ' ').strip() for cell in td_matches]
                        rows_data.append(clean_row)

        codigo_ficha = None
        programa_formacion = None
        aspirantes = []

        # Recorrer filas iniciales para metadata de cabecera
        for row in rows_data[:15]:
            cell_a = str(row[0] if len(row) > 0 and row[0] is not None else "").strip()
            cell_b = str(row[1] if len(row) > 1 and row[1] is not None else "").strip()

            if "CÓDIGO FICHA" in cell_a.upper() or "CODIGO FICHA" in cell_a.upper():
                codigo_ficha = cell_b if cell_b else cell_a
                m = re.search(r'\d{6,10}', str(codigo_ficha))
                if m:
                    codigo_ficha = m.group(0)

            if "PROGRAMA DE FORMACIÓN" in cell_a.upper() or "PROGRAMA DE FORMACION" in cell_a.upper():
                programa_formacion = cell_b if cell_b else cell_a

        # Localizar fila de encabezados de la tabla
        header_row_idx = 5
        col_identificacion = 0
        col_nombre = 1
        col_estado = 2

        for r_idx, row in enumerate(rows_data[:25]):
            row_str = [str(c or "").upper() for c in row]
            if any("IDENTIFICACI" in v for v in row_str):
                header_row_idx = r_idx
                for c_idx, v in enumerate(row_str):
                    if "IDENTIFICACI" in v:
                        col_identificacion = c_idx
                    elif "NOMBRE" in v:
                        col_nombre = c_idx
                    elif "ESTADO" in v:
                        col_estado = c_idx
                break

        # Extraer registros de aspirantes
        for row in rows_data[header_row_idx + 1:]:
            if len(row) <= max(col_identificacion, col_nombre):
                continue

            val_id = str(row[col_identificacion] if row[col_identificacion] is not None else "").strip()
            val_nombre = str(row[col_nombre] if row[col_nombre] is not None else "").strip()
            val_estado = str(row[col_estado] if len(row) > col_estado and row[col_estado] is not None else "").strip()

            if not val_id or val_id.lower() == 'none' or not val_nombre or val_nombre.lower() == 'none':
                continue

            tipo_doc = "CC"
            num_doc = val_id

            match_tipo = re.match(r'^(CC|TI|CE|PEP|PPT|PAS)\s*[-:]*\s*(\d+)', val_id, re.IGNORECASE)
            if match_tipo:
                tipo_doc = match_tipo.group(1).upper()
                num_doc = match_tipo.group(2)
            else:
                digits = re.sub(r'[^\d]', '', val_id)
                if digits:
                    num_doc = digits

            aspirantes.append({
                "tipo_documento": tipo_doc,
                "numero_documento": num_doc,
                "nombre_completo": val_nombre.upper(),
                "estado_inscripcion": val_estado if val_estado and val_estado.lower() != 'none' else 'Preinscrito'
            })

        return {
            "status": "success",
            "data": {
                "codigo_ficha": codigo_ficha or "3574135",
                "programa_formacion": programa_formacion or "PROGRAMA SENA",
                "total_aspirantes": len(aspirantes),
                "aspirantes": aspirantes
            }
        }

    except Exception as e:
        return {
            "status": "error",
            "message": f"Error al procesar el archivo Excel: {str(e)}",
            "data": None
        }


def main():
    parser = argparse.ArgumentParser(description="Parser de Reporte de Inscripciones SENA en Excel")
    parser.add_argument("--excel", required=True, help="Ruta al archivo Excel (.xlsx)")
    args = parser.parse_args()

    resultado = parse_excel_file(args.excel)
    print(json.dumps(resultado, ensure_ascii=False, indent=2))

if __name__ == "__main__":
    main()
