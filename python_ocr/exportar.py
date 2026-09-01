"""Genera reportes comparativos profesionales en Excel (XLSX) y CSV."""
import csv
import io
import os

from campos import calcular_edad
from cotejo import CAMPOS

# Como se llama cada estado en el archivo que se entrega.
NOMBRE_ESTADO = {
    "ok": "Correcto",
    "revisar": "Con discrepancia",
    "sin_listado": "Solo en PDF"
}

def a_csv(personas, faltantes=()):
    salida = io.StringIO()
    w = csv.writer(salida, delimiter=";", lineterminator="\n")
    w.writerow([
        "Pág", "Tipo (PDF)", "Documento (PDF)", "Nombres (PDF)", "Apellidos (PDF)",
        "Tipo (Excel)", "Documento (Excel)", "Nombre Completo (Excel)",
        "Estado Cotejo", "Novedad / Observación", "Fecha Nacimiento", "RH", "Edad", "Corregido a Mano"
    ])
    for per in personas:
        vals = per.get("valores", {})
        ref = per.get("referencia") or per.get("listado") or {}
        pags = ", ".join(str(p.get("pagina", "")) for p in per.get("paginas", [])) or "1"
        edad = calcular_edad(vals.get("nacimiento", ""))
        
        w.writerow([
            pags,
            vals.get("tipo_documento", "CC"),
            vals.get("documento", ""),
            vals.get("nombres", ""),
            vals.get("apellidos", ""),
            ref.get("tipo", ref.get("tipo_documento", "")),
            ref.get("documento", ""),
            ref.get("nombre_completo", ""),
            NOMBRE_ESTADO.get(per.get("estado"), per.get("estado", "")),
            per.get("novedad", per.get("estado_texto", "")),
            vals.get("nacimiento", ""),
            vals.get("rh", ""),
            "" if edad is None else edad,
            "Sí" if per.get("editado") else "No"
        ])
    if faltantes:
        w.writerow([])
        w.writerow(["Solo en Excel: aspirantes del listado sin cédula en el PDF"])
        w.writerow(["Tipo (Excel)", "Documento (Excel)", "Nombre Completo (Excel)", "Estado"])
        for f in faltantes:
            w.writerow([f.get("tipo", "CC"), f.get("documento", ""), f.get("nombre_completo", ""), "Solo en Excel (Sin Cédula)"])
    return salida.getvalue().encode("utf-8-sig")


def a_excel(personas, faltantes=(), ficha_info=None):
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    
    # ------------------ ESTILOS ------------------
    fuente_titulo = Font(name="Segoe UI", size=14, bold=True, color="1E293B")
    fuente_subtitulo = Font(name="Segoe UI", size=10, italic=True, color="64748B")
    fuente_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    fuente_normal = Font(name="Segoe UI", size=9, color="0F172A")
    fuente_bold = Font(name="Segoe UI", size=9, bold=True, color="0F172A")
    
    fill_header_pdf = PatternFill("solid", fgColor="1E40AF")    # Azul Oscuro
    fill_header_excel = PatternFill("solid", fgColor="047857")  # Verde Oscuro
    fill_header_res = PatternFill("solid", fgColor="334155")    # Gris Oscuro
    
    fill_ok = PatternFill("solid", fgColor="DCFCE7")       # Verde claro
    fill_revisar = PatternFill("solid", fgColor="FEF9C3")  # Amarillo claro
    fill_solo_pdf = PatternFill("solid", fgColor="FEE2E2") # Rojo claro
    fill_solo_exc = PatternFill("solid", fgColor="F1F5F9") # Gris claro

    borde_fino = Side(style='thin', color="CBD5E1")
    borde_celda = Border(left=borde_fino, right=borde_fino, top=borde_fino, bottom=borde_fino)
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")

    # ================== HOJA 1: COMPARATIVA COMPLETA ==================
    ws = wb.active
    ws.title = "Cruce PDF vs Excel"
    ws.views.sheetView[0].showGridLines = True

    # Encabezado Informativo
    cod_ficha = (ficha_info or {}).get("codigo_ficha", "")
    prog_ficha = (ficha_info or {}).get("programa_formacion", "Reporte de Cotejo y Validación Documental")
    
    ws.merge_cells("A1:M1")
    ws["A1"] = f"INFORME DE COTEJO Y CRUCE DOCUMENTAL {f'- FICHA {cod_ficha}' if cod_ficha else ''}"
    ws["A1"].font = fuente_titulo
    ws["A1"].alignment = align_left

    ws.merge_cells("A2:M2")
    ws["A2"] = f"Programa: {prog_ficha} | Generado por Sistema OCR & Cotejo Automático"
    ws["A2"].font = fuente_subtitulo
    ws["A2"].alignment = align_left

    # Encabezados de Columnas agrupados
    headers = [
        ("Pág", fill_header_pdf),
        ("Tipo (PDF)", fill_header_pdf),
        ("Documento (PDF)", fill_header_pdf),
        ("Nombres Extraídos (PDF)", fill_header_pdf),
        ("Apellidos Extraídos (PDF)", fill_header_pdf),
        ("Tipo (Excel)", fill_header_excel),
        ("Documento (Excel)", fill_header_excel),
        ("Nombre Completo (Excel)", fill_header_excel),
        ("Estado de Cotejo", fill_header_res),
        ("Novedad / Observación", fill_header_res),
        ("F. Nacimiento", fill_header_res),
        ("RH", fill_header_res),
        ("Edad", fill_header_res),
        ("Corregido a Mano", fill_header_res)
    ]

    fila_h = 4
    for col_idx, (h_texto, h_fill) in enumerate(headers, 1):
        c = ws.cell(row=fila_h, column=col_idx, value=h_texto)
        c.font = fuente_header
        c.fill = h_fill
        c.alignment = align_center
        c.border = borde_celda

    # Filas de datos
    fila_actual = 5
    for per in personas:
        vals = per.get("valores", {})
        ref = per.get("referencia") or per.get("listado") or {}
        pags = ", ".join(str(p.get("pagina", "")) for p in per.get("paginas", [])) or "1"
        edad = calcular_edad(vals.get("nacimiento", ""))
        estado = per.get("estado", "ok")
        
        fill_estado = fill_ok if estado == 'ok' else (fill_revisar if estado == 'revisar' else fill_solo_pdf)
        
        datos_fila = [
            (pags, align_center),
            (vals.get("tipo_documento", "CC"), align_center),
            (vals.get("documento", ""), align_left),
            (vals.get("nombres", ""), align_left),
            (vals.get("apellidos", ""), align_left),
            (ref.get("tipo", ref.get("tipo_documento", "-")), align_center),
            (ref.get("documento", "-"), align_left),
            (ref.get("nombre_completo", "-"), align_left),
            (NOMBRE_ESTADO.get(estado, estado), align_center),
            (per.get("novedad", per.get("estado_texto", "Correcto")), align_left),
            (vals.get("nacimiento", "-"), align_center),
            (vals.get("rh", "-"), align_center),
            (edad if edad is not None else "-", align_center),
            ("Sí" if per.get("editado") else "No", align_center)
        ]

        for col_idx, (valor, alineacion) in enumerate(datos_fila, 1):
            celda = ws.cell(row=fila_actual, column=col_idx, value=valor)
            celda.font = fuente_normal
            celda.alignment = alineacion
            celda.border = borde_celda
            if col_idx == 9: # Columna Estado
                celda.fill = fill_estado
                celda.font = fuente_bold

        fila_actual += 1

    # Agregar también las faltantes (Solo en Excel) al final de la matriz para vista integral
    for f in faltantes:
        datos_faltante = [
            ("-", align_center),
            ("-", align_center),
            ("-", align_center),
            ("-", align_center),
            ("-", align_center),
            (f.get("tipo", "CC"), align_center),
            (f.get("documento", ""), align_left),
            (f.get("nombre_completo", ""), align_left),
            ("Solo en Excel", align_center),
            ("No adjuntó cédula en el PDF", align_left),
            ("-", align_center),
            ("-", align_center),
            ("-", align_center),
            ("No", align_center)
        ]
        for col_idx, (valor, alineacion) in enumerate(datos_faltante, 1):
            celda = ws.cell(row=fila_actual, column=col_idx, value=valor)
            celda.font = fuente_normal
            celda.alignment = alineacion
            celda.border = borde_celda
            celda.fill = fill_solo_exc
            if col_idx == 9:
                celda.font = fuente_bold
        fila_actual += 1

    # Ajuste automático de anchos de columna (evaluando solo a partir de la fila 4 de encabezados y datos)
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            if cell.row >= 4:  # Ignorar títulos combinados de filas 1 y 2
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 8)

    ws.freeze_panes = "A5"

    # ================== HOJA 2: SOLO EN EXCEL (FALTANTES) ==================
    if faltantes:
        ws2 = wb.create_sheet(title="Solo en Excel")
        ws2.views.sheetView[0].showGridLines = True
        
        ws2["A1"] = "ASPIRANTES EN LISTADO EXCEL SIN CÉDULA EN PDF"
        ws2["A1"].font = fuente_titulo
        
        headers2 = ["#", "Tipo Doc", "Número de Documento", "Nombre Completo", "Estado"]
        for col_idx, h in enumerate(headers2, 1):
            c = ws2.cell(row=3, column=col_idx, value=h)
            c.font = fuente_header
            c.fill = fill_header_excel
            c.alignment = align_center
            c.border = borde_celda
            
        for idx, f in enumerate(faltantes, 1):
            row_num = idx + 3
            ws2.cell(row=row_num, column=1, value=idx).alignment = align_center
            ws2.cell(row=row_num, column=2, value=f.get("tipo", "CC")).alignment = align_center
            ws2.cell(row=row_num, column=3, value=f.get("documento", "")).alignment = align_left
            ws2.cell(row=row_num, column=4, value=f.get("nombre_completo", "")).alignment = align_left
            ws2.cell(row=row_num, column=5, value="Sin Cédula en PDF").alignment = align_center
            
            for c_idx in range(1, 6):
                cel = ws2.cell(row=row_num, column=c_idx)
                cel.font = fuente_normal
                cel.border = borde_celda

        for col in ws2.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws2.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 4, 12)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
